# activities/excel_service.py
"""
Excel export/import service for Activity Sheets.
Supports streaming for large files and validation on import.
"""

import io
from typing import List, Dict, Any, Generator, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.conf import settings

from .models import ActivitySheet, ActivitySheetRow, ActivityTemplate


class ExcelService:
    """Service for Excel export/import operations"""
    
    # Style constants
    HEADER_FILL = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
    BORDER = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    ALT_ROW_FILL = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    
    def __init__(self, sheet: ActivitySheet):
        self.sheet = sheet
        self.columns = sheet.column_snapshot or []
    
    def export_to_excel(self, include_data: bool = True) -> io.BytesIO:
        """
        Export sheet to Excel file.
        Returns BytesIO buffer with Excel content.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = 'الأنشطة'
        ws.sheet_view.rightToLeft = True
        
        # Get column count
        col_count = len(self.columns)
        if col_count == 0:
            col_count = 1  # At least one column
        
        # Row 1: Header placeholder (can add image later)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        ws.row_dimensions[1].height = 70
        header_cell = ws.cell(row=1, column=1)
        header_cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        header_cell.value = self.sheet.name if self.sheet.name else 'ورقة الأنشطة'
        header_cell.font = Font(bold=True, size=16, color='1E3A5F')
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Row 2: Column headers
        self._add_column_headers(ws)
        
        # Row 3+: Data rows
        if include_data:
            self._add_data_rows(ws)
        else:
            # Add 10 empty rows for template
            self._add_empty_rows(ws, 10)
        
        # Set column widths
        self._set_column_widths(ws)
        
        # Add AutoFilter on header row
        if col_count > 0:
            ws.auto_filter.ref = f'A2:{get_column_letter(col_count)}2'
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def _add_column_headers(self, ws):
        """Add column headers to row 2"""
        for col_idx, col_def in enumerate(self.columns, start=1):
            cell = ws.cell(row=2, column=col_idx)
            # Get label from snapshot
            label = col_def.get('label') or col_def.get('column_definition__label', f'Column {col_idx}')
            cell.value = label
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.BORDER
        
        ws.row_dimensions[2].height = 40
    
    def _add_data_rows(self, ws):
        """Add data rows starting from row 3"""
        rows = self.sheet.rows.order_by('row_number')
        
        for row_idx, row in enumerate(rows, start=3):
            for col_idx, col_def in enumerate(self.columns, start=1):
                col_key = col_def.get('key') or col_def.get('column_definition__key', '')
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = row.data.get(col_key, '')
                cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
                cell.border = self.BORDER
                
                # Alternate row background
                if row_idx % 2 == 1:
                    cell.fill = self.ALT_ROW_FILL
                
                # Apply styles from row
                style = (row.styles or {}).get(col_key, {})
                if style:
                    self._apply_cell_style(cell, style, row_idx)
            
            # Set row height
            row_height = getattr(row, 'height', 32) or 32
            ws.row_dimensions[row_idx].height = row_height / 1.5
    
    def _add_empty_rows(self, ws, count: int):
        """Add empty rows for template"""
        for row_idx in range(3, 3 + count):
            for col_idx in range(1, len(self.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = ''
                cell.border = self.BORDER
                cell.alignment = Alignment(horizontal='right', vertical='center')
                
                if row_idx % 2 == 1:
                    cell.fill = self.ALT_ROW_FILL
            
            ws.row_dimensions[row_idx].height = 25
    
    def _apply_cell_style(self, cell, style: Dict, row_idx: int):
        """Apply style dict to cell"""
        font_kwargs = {}
        
        if style.get('bold'):
            font_kwargs['bold'] = True
        if style.get('italic'):
            font_kwargs['italic'] = True
        if style.get('textColor'):
            color = style['textColor'].replace('#', '')
            if len(color) == 6:
                font_kwargs['color'] = color
        
        if font_kwargs:
            cell.font = Font(**font_kwargs)
        
        if style.get('backgroundColor'):
            color = style['backgroundColor'].replace('#', '')
            if len(color) == 6:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        elif row_idx % 2 == 1:
            # Keep alternate row styling if no custom background
            cell.fill = self.ALT_ROW_FILL
    
    def _set_column_widths(self, ws):
        """Set column widths based on column definitions"""
        for col_idx, col_def in enumerate(self.columns, start=1):
            # Get width from snapshot
            width = col_def.get('width') or col_def.get('column_definition__default_width', 120)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(width / 7, 10)
    
    def import_from_excel(self, file_content: bytes, validate: bool = False) -> Dict[str, Any]:
        """
        Import Excel file to sheet.
        Returns import statistics.
        
        Args:
            file_content: Raw bytes of Excel file
            validate: Whether to validate data against column rules
        
        Returns:
            Dict with import results
        """
        try:
            wb = load_workbook(io.BytesIO(file_content))
        except Exception as e:
            raise ValueError(f'Invalid Excel file: {str(e)}')
        
        ws = wb.active
        
        if ws.max_row < 3:
            raise ValueError('Excel file must have at least a header row and one data row')
        
        # Build header map from row 2
        header_map = self._build_header_map(ws)
        
        if not header_map:
            raise ValueError('No matching columns found in Excel file. Please use the correct template.')
        
        # Collect all row data
        all_rows = []
        for row_idx in range(3, ws.max_row + 1):
            row_data = {}
            has_data = False
            
            for col_idx, col_key in header_map.items():
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None and str(cell_value).strip():
                    row_data[col_key] = str(cell_value).strip()
                    has_data = True
            
            if has_data:
                all_rows.append(row_data)
        
        if not all_rows:
            raise ValueError('No data rows found in Excel file')
        
        # Optional validation
        validation_errors = []
        if validate:
            from .validators import RowValidator
            validator = RowValidator(self.columns)
            validation_errors = validator.validate_rows(all_rows)
            
            if validation_errors:
                return {
                    'success': False,
                    'validation_errors': validation_errors,
                    'message': f'{len(validation_errors)} rows have validation errors'
                }
        
        # Import data
        imported_count = 0
        updated_count = 0
        
        for idx, row_data in enumerate(all_rows, start=1):
            obj, created = ActivitySheetRow.objects.update_or_create(
                sheet=self.sheet,
                row_number=idx,
                defaults={
                    'data': row_data,
                    'styles': {},
                    'height': 32
                }
            )
            
            if created:
                imported_count += 1
            else:
                updated_count += 1
        
        # Delete any rows beyond imported count
        ActivitySheetRow.objects.filter(
            sheet=self.sheet,
            row_number__gt=len(all_rows)
        ).delete()
        
        # Update sheet row count
        self.sheet.row_count = len(all_rows)
        self.sheet.save(update_fields=['row_count', 'updated_at'])
        
        return {
            'success': True,
            'imported': imported_count,
            'updated': updated_count,
            'total_rows': self.sheet.row_count
        }
    
    def _build_header_map(self, ws) -> Dict[int, str]:
        """
        Build mapping from Excel column index to column key.
        Matches headers in row 2 to column definitions.
        """
        header_map = {}
        
        for col_idx in range(1, ws.max_column + 1):
            header_value = ws.cell(row=2, column=col_idx).value
            if not header_value:
                continue
            
            header_normalized = str(header_value).replace('\n', ' ').strip().lower()
            
            # Find matching column definition
            for col_def in self.columns:
                label = col_def.get('label') or col_def.get('column_definition__label', '')
                label_normalized = label.replace('\n', ' ').strip().lower()
                
                # Match by label
                if label_normalized == header_normalized:
                    col_key = col_def.get('key') or col_def.get('column_definition__key', '')
                    if col_key:
                        header_map[col_idx] = col_key
                    break
                
                # Partial match (contains)
                if label_normalized and (
                    label_normalized in header_normalized or 
                    header_normalized in label_normalized
                ):
                    col_key = col_def.get('key') or col_def.get('column_definition__key', '')
                    if col_key:
                        header_map[col_idx] = col_key
                    break
        
        return header_map


def export_sheet_streaming(sheet_id: int) -> Generator[bytes, None, None]:
    """
    Generator for streaming Excel export (for large files).
    Yields chunks of Excel file content.
    
    Args:
        sheet_id: ID of the sheet to export
    
    Yields:
        Chunks of Excel file bytes
    """
    sheet = ActivitySheet.objects.get(id=sheet_id)
    service = ExcelService(sheet)
    
    buffer = service.export_to_excel(include_data=True)
    
    # Yield in chunks
    chunk_size = 8192
    while True:
        chunk = buffer.read(chunk_size)
        if not chunk:
            break
        yield chunk


def create_template_excel(template: ActivityTemplate) -> io.BytesIO:
    """
    Create an empty Excel template from a template's column configuration.
    
    Args:
        template: ActivityTemplate instance
    
    Returns:
        BytesIO buffer with Excel content
    """
    # Build column snapshot from template
    column_snapshot = list(
        template.template_columns.select_related('column_definition')
        .order_by('order')
        .values(
            'column_definition__key',
            'column_definition__label',
            'column_definition__data_type',
            'column_definition__default_width',
            'order',
            'width',
            'is_required'
        )
    )
    
    # Transform snapshot format
    columns = []
    for col in column_snapshot:
        columns.append({
            'key': col['column_definition__key'],
            'label': col['column_definition__label'],
            'data_type': col['column_definition__data_type'],
            'width': col['width'] or col['column_definition__default_width'] or 120,
            'order': col['order'],
            'is_required': col['is_required']
        })
    
    # Create a mock sheet object for the service
    class MockSheet:
        def __init__(self, name, column_snapshot):
            self.name = name
            self.column_snapshot = column_snapshot
    
    mock_sheet = MockSheet(
        name=f'{template.name} - قالب',
        column_snapshot=columns
    )
    
    service = ExcelService(mock_sheet)
    return service.export_to_excel(include_data=False)
